"""Rhisos Mobilya CRM backend tests - auth, customers, quotes, transactions, dashboard."""
import os
import time
from datetime import datetime, timezone, timedelta
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://rhisos-crm.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "iskberkan@gmail.com"
ADMIN_PASSWORD = "rhisos2026"


@pytest.fixture(scope="session")
def admin_session():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    data = r.json()
    assert data.get("email") == ADMIN_EMAIL
    assert data.get("role") == "admin"
    return s


# ---------------- Auth ----------------
class TestAuth:
    def test_login_success(self, admin_session):
        r = admin_session.get(f"{API}/auth/me")
        assert r.status_code == 200
        assert r.json()["email"] == ADMIN_EMAIL

    def test_login_invalid(self):
        r = requests.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": "wrong-pass-xxx"})
        assert r.status_code == 401

    def test_me_unauthorized(self):
        r = requests.get(f"{API}/auth/me")
        assert r.status_code == 401

    def test_register_new_user(self):
        email = f"test_user_{int(time.time())}@example.com"
        r = requests.post(f"{API}/auth/register", json={"email": email, "password": "pass1234", "name": "TEST User"})
        assert r.status_code == 200, r.text
        assert r.json()["email"] == email


# ---------------- Customers ----------------
class TestCustomers:
    created_id = None

    def test_create_customer(self, admin_session):
        payload = {"name": "TEST_Musteri", "company": "TEST Co", "email": "t@test.com", "phone": "555", "address": "adr", "notes": "n"}
        r = admin_session.post(f"{API}/customers", json=payload)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["name"] == "TEST_Musteri"
        assert "id" in data
        TestCustomers.created_id = data["id"]

    def test_get_customers(self, admin_session):
        r = admin_session.get(f"{API}/customers")
        assert r.status_code == 200
        assert any(c["id"] == TestCustomers.created_id for c in r.json())

    def test_update_customer(self, admin_session):
        payload = {"name": "TEST_Updated", "company": "X", "email": "", "phone": "", "address": "", "notes": ""}
        r = admin_session.put(f"{API}/customers/{TestCustomers.created_id}", json=payload)
        assert r.status_code == 200
        assert r.json()["name"] == "TEST_Updated"

    def test_delete_customer_last(self, admin_session):
        # keep customer for quote tests, defer delete
        pass


# ---------------- Quotes ----------------
class TestQuotes:
    customer_id = None
    quote_id = None

    def test_setup_customer(self, admin_session):
        r = admin_session.post(f"{API}/customers", json={"name": "TEST_QCust"})
        assert r.status_code == 200
        TestQuotes.customer_id = r.json()["id"]

    def test_create_quote_totals(self, admin_session):
        payload = {
            "customer_id": TestQuotes.customer_id,
            "customer_name": "TEST_QCust",
            "title": "TEST Teklif",
            "items": [
                {"description": "Koltuk", "quantity": 2, "unit_price": 1000, "vat_rate": 20},
                {"description": "Masa", "quantity": 1, "unit_price": 500, "vat_rate": 20},
            ],
            "currency": "TRY",
            "discount": 100,
        }
        r = admin_session.post(f"{API}/quotes", json=payload)
        assert r.status_code == 200, r.text
        d = r.json()
        # subtotal = 2500, vat = 500, disc 100 -> grand 2900
        assert d["subtotal"] == 2500.0
        assert d["vat_total"] == 500.0
        assert d["grand_total"] == 2900.0
        assert d["status"] == "pending"
        assert d["quote_number"].startswith("TKF-")
        TestQuotes.quote_id = d["id"]

    def test_get_quote(self, admin_session):
        r = admin_session.get(f"{API}/quotes/{TestQuotes.quote_id}")
        assert r.status_code == 200
        assert r.json()["id"] == TestQuotes.quote_id

    def test_update_quote(self, admin_session):
        payload = {
            "customer_id": TestQuotes.customer_id,
            "customer_name": "TEST_QCust",
            "title": "TEST Teklif V2",
            "items": [{"description": "Sandalye", "quantity": 4, "unit_price": 250, "vat_rate": 10}],
            "currency": "TRY",
            "discount": 0,
        }
        r = admin_session.put(f"{API}/quotes/{TestQuotes.quote_id}", json=payload)
        assert r.status_code == 200
        d = r.json()
        assert d["subtotal"] == 1000.0
        assert d["vat_total"] == 100.0
        assert d["grand_total"] == 1100.0
        assert d["title"] == "TEST Teklif V2"

    def test_approve_does_NOT_create_income(self, admin_session):
        """MODEL CHANGE: approving a quote no longer auto-creates income - income only from payments."""
        r = admin_session.patch(f"{API}/quotes/{TestQuotes.quote_id}/status", json={"status": "approved"})
        assert r.status_code == 200
        assert r.json()["status"] == "approved"
        tx = admin_session.get(f"{API}/transactions", params={"type": "income"}).json()
        auto = [t for t in tx if t.get("quote_id") == TestQuotes.quote_id and t.get("auto")]
        assert len(auto) == 0, f"Approving quote should NOT create income, found: {auto}"

    def test_revert_status(self, admin_session):
        r = admin_session.patch(f"{API}/quotes/{TestQuotes.quote_id}/status", json={"status": "pending"})
        assert r.status_code == 200

    def test_delete_quote(self, admin_session):
        r = admin_session.delete(f"{API}/quotes/{TestQuotes.quote_id}")
        assert r.status_code == 200
        # verify quote gone
        r2 = admin_session.get(f"{API}/quotes/{TestQuotes.quote_id}")
        assert r2.status_code == 404
        # verify auto txn removed
        tx = admin_session.get(f"{API}/transactions").json()
        auto = [t for t in tx if t.get("quote_id") == TestQuotes.quote_id]
        assert len(auto) == 0
        # cleanup customer
        admin_session.delete(f"{API}/customers/{TestQuotes.customer_id}")


# ---------------- Transactions ----------------
class TestTransactions:
    tid = None

    def test_create_income(self, admin_session):
        r = admin_session.post(f"{API}/transactions", json={
            "type": "income", "amount": 500, "category": "TEST_Satis",
            "payment_method": "cash", "date": "2026-01-10", "currency": "TRY",
        })
        assert r.status_code == 200
        TestTransactions.tid = r.json()["id"]

    def test_create_expense(self, admin_session):
        r = admin_session.post(f"{API}/transactions", json={
            "type": "expense", "amount": 200, "category": "TEST_Kira",
            "payment_method": "bank", "date": "2026-01-11", "currency": "TRY",
        })
        assert r.status_code == 200

    def test_filter_type(self, admin_session):
        r = admin_session.get(f"{API}/transactions", params={"type": "income"})
        assert r.status_code == 200
        assert all(t["type"] == "income" for t in r.json())

    def test_filter_date(self, admin_session):
        r = admin_session.get(f"{API}/transactions", params={"start": "2026-01-10", "end": "2026-01-10"})
        assert r.status_code == 200
        for t in r.json():
            assert "2026-01-10" <= t["date"] <= "2026-01-10"

    def test_delete_transaction(self, admin_session):
        r = admin_session.delete(f"{API}/transactions/{TestTransactions.tid}")
        assert r.status_code == 200


# ---------------- Dashboard ----------------
class TestDashboard:
    def test_stats(self, admin_session):
        r = admin_session.get(f"{API}/dashboard/stats")
        assert r.status_code == 200
        d = r.json()
        for k in ["total_income", "total_expense", "balance", "monthly", "expense_by_category", "quote_counts", "customer_count", "expiring_quotes"]:
            assert k in d
        assert isinstance(d["expiring_quotes"], list)


# ---------------- Settings (NEW) ----------------
class TestSettings:
    def test_get_default(self, admin_session):
        r = admin_session.get(f"{API}/settings")
        assert r.status_code == 200
        d = r.json()
        for k in ["company_name", "tagline", "address", "phone", "email", "website", "tax_office", "tax_number", "logo"]:
            assert k in d

    def test_update_and_persist(self, admin_session):
        payload = {
            "company_name": "TEST Rhisos Mobilya",
            "tagline": "TEST Özel Tasarım",
            "address": "TEST Adres 123",
            "phone": "+90 555 111 2233",
            "email": "test@rhisos.com",
            "website": "https://rhisos.test",
            "tax_office": "TEST Vergi Dairesi",
            "tax_number": "1234567890",
            "logo": "",
        }
        r = admin_session.put(f"{API}/settings", json=payload)
        assert r.status_code == 200, r.text
        # GET to verify persistence
        r2 = admin_session.get(f"{API}/settings")
        assert r2.status_code == 200
        d = r2.json()
        assert d["company_name"] == payload["company_name"]
        assert d["phone"] == payload["phone"]
        assert d["email"] == payload["email"]
        assert d["tax_number"] == payload["tax_number"]


# ---------------- Expiring quotes on dashboard (NEW) ----------------
class TestExpiringQuotes:
    customer_id = None
    expiring_qid = None
    far_qid = None

    def test_setup(self, admin_session):
        r = admin_session.post(f"{API}/customers", json={"name": "TEST_ExpCust"})
        assert r.status_code == 200
        TestExpiringQuotes.customer_id = r.json()["id"]

        soon = (datetime.now(timezone.utc).date() + timedelta(days=3)).isoformat()
        far = (datetime.now(timezone.utc).date() + timedelta(days=60)).isoformat()

        for name, vu, target in [("soon", soon, "expiring_qid"), ("far", far, "far_qid")]:
            payload = {
                "customer_id": TestExpiringQuotes.customer_id,
                "customer_name": "TEST_ExpCust",
                "title": f"TEST Expiring {name}",
                "items": [{"description": "x", "quantity": 1, "unit_price": 100, "vat_rate": 20}],
                "currency": "TRY",
                "valid_until": vu,
                "discount": 0,
            }
            r = admin_session.post(f"{API}/quotes", json=payload)
            assert r.status_code == 200
            setattr(TestExpiringQuotes, target, r.json()["id"])

    def test_dashboard_lists_expiring(self, admin_session):
        r = admin_session.get(f"{API}/dashboard/stats")
        assert r.status_code == 200
        ids = [q["id"] for q in r.json()["expiring_quotes"]]
        assert TestExpiringQuotes.expiring_qid in ids
        assert TestExpiringQuotes.far_qid not in ids

    def test_cleanup(self, admin_session):
        admin_session.delete(f"{API}/quotes/{TestExpiringQuotes.expiring_qid}")
        admin_session.delete(f"{API}/quotes/{TestExpiringQuotes.far_qid}")
        admin_session.delete(f"{API}/customers/{TestExpiringQuotes.customer_id}")


# ---------------- Excel Export (NEW) ----------------
class TestExcelExport:
    XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    tids = []

    def test_setup(self, admin_session):
        for t, cat, amt, date_str in [
            ("income", "TEST_ExportIn", 1234.5, "2026-01-05"),
            ("expense", "TEST_ExportEx", 250, "2026-01-06"),
        ]:
            r = admin_session.post(f"{API}/transactions", json={
                "type": t, "amount": amt, "category": cat,
                "payment_method": "bank", "date": date_str, "currency": "TRY",
                "description": f"TEST xlsx {t}",
            })
            assert r.status_code == 200
            TestExcelExport.tids.append(r.json()["id"])

    def test_export_all(self, admin_session):
        r = admin_session.get(f"{API}/transactions/export")
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith(self.XLSX_MIME)
        assert len(r.content) > 500  # non-empty xlsx
        # PK zip header (xlsx = zip)
        assert r.content[:2] == b"PK"

    def test_export_type_filter(self, admin_session):
        r = admin_session.get(f"{API}/transactions/export", params={"type": "expense"})
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith(self.XLSX_MIME)
        # Parse workbook and ensure only expense rows appear
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # data rows before totals: type column index 1 (0-based)
        data_types = [row[1] for row in rows if row[0]]  # skip totals (empty tarih)
        assert data_types, "Export contained no data rows"
        assert all(t == "Gider" for t in data_types)

    def test_export_date_filter(self, admin_session):
        r = admin_session.get(f"{API}/transactions/export", params={"start": "2026-01-05", "end": "2026-01-05"})
        assert r.status_code == 200
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        dates = [row[0] for row in rows if row[0]]
        assert dates and all(d == "2026-01-05" for d in dates)

    def test_cleanup(self, admin_session):
        for tid in TestExcelExport.tids:
            admin_session.delete(f"{API}/transactions/{tid}")


# ---------------- Kapora / Payments (NEW iter3) ----------------
class TestQuotePayments:
    customer_id = None
    quote_id = None
    payment_id = None

    def test_setup(self, admin_session):
        r = admin_session.post(f"{API}/customers", json={"name": "TEST_PayCust", "email": ""})
        assert r.status_code == 200
        TestQuotePayments.customer_id = r.json()["id"]
        payload = {
            "customer_id": TestQuotePayments.customer_id,
            "customer_name": "TEST_PayCust",
            "title": "TEST Pay Teklif",
            "items": [{"description": "Set", "quantity": 1, "unit_price": 10000, "vat_rate": 20}],
            "currency": "TRY", "discount": 0,
        }
        r = admin_session.post(f"{API}/quotes", json=payload)
        assert r.status_code == 200
        d = r.json()
        assert d["grand_total"] == 12000.0
        assert d["paid_total"] == 0
        assert d.get("payments") == []
        TestQuotePayments.quote_id = d["id"]

    def test_add_payment_creates_income_txn(self, admin_session):
        r = admin_session.post(f"{API}/quotes/{TestQuotePayments.quote_id}/payments", json={
            "amount": 5000, "date": "2026-01-15", "method": "bank", "note": "TEST kapora"
        })
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["paid_total"] == 5000.0
        assert len(d["payments"]) == 1
        TestQuotePayments.payment_id = d["payments"][0]["id"]
        # verify linked cash flow txn
        tx = admin_session.get(f"{API}/transactions", params={"type": "income"}).json()
        linked = [t for t in tx if t.get("quote_id") == TestQuotePayments.quote_id]
        assert len(linked) == 1
        assert linked[0]["amount"] == 5000.0
        assert linked[0]["category"] == "Kapora / Teklif Ödemesi"
        assert linked[0].get("payment_id") == TestQuotePayments.payment_id
        assert linked[0].get("auto") is True

    def test_remaining_balance(self, admin_session):
        r = admin_session.get(f"{API}/quotes/{TestQuotePayments.quote_id}")
        assert r.status_code == 200
        d = r.json()
        assert d["grand_total"] - d["paid_total"] == 7000.0

    def test_reject_zero_or_negative_amount(self, admin_session):
        r = admin_session.post(f"{API}/quotes/{TestQuotePayments.quote_id}/payments",
                               json={"amount": 0, "date": "2026-01-15", "method": "cash"})
        assert r.status_code == 400

    def test_delete_payment_removes_txn(self, admin_session):
        r = admin_session.delete(f"{API}/quotes/{TestQuotePayments.quote_id}/payments/{TestQuotePayments.payment_id}")
        assert r.status_code == 200
        d = r.json()
        assert d["paid_total"] == 0
        assert d["payments"] == []
        tx = admin_session.get(f"{API}/transactions").json()
        linked = [t for t in tx if t.get("payment_id") == TestQuotePayments.payment_id]
        assert len(linked) == 0

    def test_cleanup(self, admin_session):
        admin_session.delete(f"{API}/quotes/{TestQuotePayments.quote_id}")
        admin_session.delete(f"{API}/customers/{TestQuotePayments.customer_id}")


# ---------------- Monthly profit (NEW iter3) ----------------
class TestMonthlyProfit:
    tids = []

    def test_setup(self, admin_session):
        for t, cat, amt in [("income", "TEST_MPIn", 1500), ("expense", "TEST_MPEx", 400)]:
            r = admin_session.post(f"{API}/transactions", json={
                "type": t, "amount": amt, "category": cat,
                "payment_method": "cash", "date": "2026-01-20", "currency": "TRY",
            })
            assert r.status_code == 200
            TestMonthlyProfit.tids.append(r.json()["id"])

    def test_monthly_profit_current(self, admin_session):
        r = admin_session.get(f"{API}/dashboard/monthly-profit", params={"month": "2026-01"})
        assert r.status_code == 200
        d = r.json()
        assert d["month"] == "2026-01"
        assert d["income"] >= 1500
        assert d["expense"] >= 400
        assert d["profit"] == round(d["income"] - d["expense"], 2)

    def test_monthly_profit_empty_month(self, admin_session):
        r = admin_session.get(f"{API}/dashboard/monthly-profit", params={"month": "2019-01"})
        assert r.status_code == 200
        d = r.json()
        assert d["income"] == 0
        assert d["expense"] == 0
        assert d["profit"] == 0

    def test_cleanup(self, admin_session):
        for tid in TestMonthlyProfit.tids:
            admin_session.delete(f"{API}/transactions/{tid}")


# ---------------- Customer History (NEW iter3) ----------------
class TestCustomerHistory:
    customer_id = None
    quote_id = None

    def test_setup(self, admin_session):
        r = admin_session.post(f"{API}/customers", json={"name": "TEST_HistCust", "email": "hist@test.com"})
        assert r.status_code == 200
        TestCustomerHistory.customer_id = r.json()["id"]
        payload = {
            "customer_id": TestCustomerHistory.customer_id,
            "customer_name": "TEST_HistCust",
            "title": "TEST Hist Teklif",
            "items": [{"description": "Ürün", "quantity": 1, "unit_price": 1000, "vat_rate": 20}],
            "currency": "TRY", "discount": 0,
        }
        r = admin_session.post(f"{API}/quotes", json=payload)
        assert r.status_code == 200
        TestCustomerHistory.quote_id = r.json()["id"]
        # add a payment
        r = admin_session.post(f"{API}/quotes/{TestCustomerHistory.quote_id}/payments",
                               json={"amount": 300, "date": "2026-01-15", "method": "cash"})
        assert r.status_code == 200

    def test_history_endpoint(self, admin_session):
        r = admin_session.get(f"{API}/customers/{TestCustomerHistory.customer_id}/history")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["customer"]["id"] == TestCustomerHistory.customer_id
        assert d["customer"]["name"] == "TEST_HistCust"
        assert len(d["quotes"]) == 1
        assert d["quotes"][0]["id"] == TestCustomerHistory.quote_id
        assert len(d["payments"]) == 1
        assert d["payments"][0]["amount"] == 300
        assert d["payments"][0]["quote_id"] == TestCustomerHistory.quote_id
        assert d["totals"]["quote_count"] == 1
        assert d["totals"]["total_paid"] == 300
        assert d["totals"]["total_quoted"] == 1200.0  # 1000 + 20% vat

    def test_history_404(self, admin_session):
        # use a valid ObjectId format that doesn't exist
        r = admin_session.get(f"{API}/customers/507f1f77bcf86cd799439011/history")
        assert r.status_code == 404

    def test_cleanup(self, admin_session):
        admin_session.delete(f"{API}/quotes/{TestCustomerHistory.quote_id}")
        admin_session.delete(f"{API}/customers/{TestCustomerHistory.customer_id}")


# ---------------- Email Quote (NEW iter3) ----------------
class TestEmailQuote:
    customer_with_email = None
    customer_no_email = None
    quote_with_email = None
    quote_no_email = None

    def _mk_quote(self, admin_session, customer_id, customer_name):
        payload = {
            "customer_id": customer_id, "customer_name": customer_name,
            "title": "TEST Email Teklif",
            "items": [{"description": "Item", "quantity": 1, "unit_price": 500, "vat_rate": 20}],
            "currency": "TRY", "discount": 0,
        }
        r = admin_session.post(f"{API}/quotes", json=payload)
        assert r.status_code == 200
        return r.json()["id"]

    def test_setup(self, admin_session):
        r = admin_session.post(f"{API}/customers", json={"name": "TEST_EmailCust", "email": "delivered@resend.dev"})
        assert r.status_code == 200
        TestEmailQuote.customer_with_email = r.json()["id"]
        r = admin_session.post(f"{API}/customers", json={"name": "TEST_NoEmailCust", "email": ""})
        assert r.status_code == 200
        TestEmailQuote.customer_no_email = r.json()["id"]
        TestEmailQuote.quote_with_email = self._mk_quote(admin_session, TestEmailQuote.customer_with_email, "TEST_EmailCust")
        TestEmailQuote.quote_no_email = self._mk_quote(admin_session, TestEmailQuote.customer_no_email, "TEST_NoEmailCust")

    def test_email_no_customer_email_returns_400(self, admin_session):
        r = admin_session.post(f"{API}/quotes/{TestEmailQuote.quote_no_email}/email")
        assert r.status_code == 400
        assert "e-posta" in r.json().get("detail", "").lower()

    def test_email_success(self, admin_session):
        r = admin_session.post(f"{API}/quotes/{TestEmailQuote.quote_with_email}/email")
        # if managed email is not configured in this env, allow 502 gracefully but flag it
        if r.status_code == 502:
            pytest.skip("Emergent managed email service returned 502 - external dependency")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["ok"] is True
        assert d["to"] == "delivered@resend.dev"
        assert "email_id" in d

    def test_cleanup(self, admin_session):
        admin_session.delete(f"{API}/quotes/{TestEmailQuote.quote_with_email}")
        admin_session.delete(f"{API}/quotes/{TestEmailQuote.quote_no_email}")
        admin_session.delete(f"{API}/customers/{TestEmailQuote.customer_with_email}")
        admin_session.delete(f"{API}/customers/{TestEmailQuote.customer_no_email}")

