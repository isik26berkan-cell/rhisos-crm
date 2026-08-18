from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
import io
import logging
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Literal
from urllib.parse import quote_plus, urlparse
import uuid
import re
import ipaddress
import httpx
from html import escape
from html.parser import HTMLParser

import bcrypt
import jwt

from pydantic import BaseModel, EmailStr
from sqlalchemy import Column, String, Float, Integer, Boolean, Text, select, func, delete as sa_delete
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession
from sqlalchemy.orm import declarative_base
from sqlalchemy.dialects.mysql import LONGTEXT

# ---------------- Database ----------------
DB_HOST = os.environ["DB_HOST"]
DB_PORT = os.environ["DB_PORT"]
DB_NAME = os.environ["DB_NAME"]
DB_USER = os.environ["DB_USER"]
DB_PASSWORD = os.environ["DB_PASSWORD"]

DATABASE_URL = (
    f"mysql+aiomysql://{DB_USER}:{quote_plus(DB_PASSWORD)}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4"
)

engine = create_async_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=3600, echo=False)
AsyncSessionLocal = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)

Base = declarative_base()

async def get_db():
    async with AsyncSessionLocal() as session:
        yield session

def new_id() -> str:
    return str(uuid.uuid4())

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

# ---------------- ORM Models ----------------
class User(Base):
    __tablename__ = "users"
    id = Column(String(36), primary_key=True, default=new_id)
    email = Column(String(255), unique=True, index=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    name = Column(String(255))
    role = Column(String(50), default="user")
    created_at = Column(String(50))

class Customer(Base):
    __tablename__ = "customers"
    id = Column(String(36), primary_key=True, default=new_id)
    name = Column(String(255), nullable=False)
    company = Column(String(255), default="")
    email = Column(String(255), default="")
    phone = Column(String(100), default="")
    address = Column(Text)
    notes = Column(Text)
    created_at = Column(String(50))

class Quote(Base):
    __tablename__ = "quotes"
    id = Column(String(36), primary_key=True, default=new_id)
    quote_number = Column(String(50))
    customer_id = Column(String(36), index=True)
    customer_name = Column(String(255))
    title = Column(String(255))
    currency = Column(String(10), default="TRY")
    notes = Column(Text)
    valid_until = Column(String(20), default="")
    status = Column(String(20), default="pending")
    subtotal = Column(Float, default=0)
    vat_total = Column(Float, default=0)
    discount = Column(Float, default=0)
    grand_total = Column(Float, default=0)
    paid_total = Column(Float, default=0)
    created_at = Column(String(50))
    created_by = Column(String(255))
    emailed_at = Column(String(50), nullable=True)
    emailed_to = Column(String(255), nullable=True)

class QuoteItem(Base):
    __tablename__ = "quote_items"
    id = Column(String(36), primary_key=True, default=new_id)
    quote_id = Column(String(36), index=True, nullable=False)
    description = Column(Text)
    quantity = Column(Float, default=1)
    unit_price = Column(Float, default=0)
    vat_rate = Column(Float, default=20)
    position = Column(Integer, default=0)

class Payment(Base):
    __tablename__ = "payments"
    id = Column(String(36), primary_key=True, default=new_id)
    quote_id = Column(String(36), index=True, nullable=False)
    amount = Column(Float, default=0)
    date = Column(String(20))
    method = Column(String(20), default="cash")
    note = Column(Text)

class Transaction(Base):
    __tablename__ = "transactions"
    id = Column(String(36), primary_key=True, default=new_id)
    type = Column(String(20), index=True)
    amount = Column(Float, default=0)
    category = Column(String(255))
    payment_method = Column(String(20), default="cash")
    description = Column(Text)
    date = Column(String(20), index=True)
    currency = Column(String(10), default="TRY")
    quote_id = Column(String(36), nullable=True, index=True)
    payment_id = Column(String(36), nullable=True, index=True)
    auto = Column(Boolean, default=False)
    created_at = Column(String(50))

class Settings(Base):
    __tablename__ = "settings"
    key = Column(String(50), primary_key=True)
    company_name = Column(String(255), default="Rhisos Mobilya")
    tagline = Column(String(255), default="")
    address = Column(Text)
    phone = Column(String(100), default="")
    email = Column(String(255), default="")
    website = Column(String(255), default="")
    tax_office = Column(String(255), default="")
    tax_number = Column(String(100), default="")
    logo = Column(LONGTEXT)

class LoginAttempt(Base):
    __tablename__ = "login_attempts"
    identifier = Column(String(255), primary_key=True)
    count = Column(Integer, default=0)
    locked_until = Column(String(50), nullable=True)

# ---------------- App ----------------
app = FastAPI()
api_router = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"

# ---------------- Auth helpers ----------------
def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=60), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True, samesite="none", max_age=604800, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Oturum açılmadı")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Geçersiz token tipi")
        async with AsyncSessionLocal() as db:
            user = await db.get(User, payload["sub"])
            if not user:
                raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
            return {"id": user.id, "email": user.email, "name": user.name, "role": user.role, "created_at": user.created_at}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Oturum süresi doldu")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz token")

# ---------------- Pydantic Schemas ----------------
class RegisterIn(BaseModel):
    email: EmailStr
    password: str
    name: str

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class CustomerIn(BaseModel):
    name: str
    company: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    notes: Optional[str] = ""

class LineItem(BaseModel):
    description: str
    quantity: float = 1
    unit_price: float = 0
    vat_rate: float = 20

class QuoteIn(BaseModel):
    customer_id: str
    customer_name: str
    title: str = "Teklif"
    items: List[LineItem]
    currency: str = "TRY"
    notes: Optional[str] = ""
    valid_until: Optional[str] = ""
    discount: float = 0

class QuoteStatusIn(BaseModel):
    status: Literal["pending", "approved", "rejected"]

class TransactionIn(BaseModel):
    type: Literal["income", "expense"]
    amount: float
    category: str
    payment_method: str = "cash"
    description: Optional[str] = ""
    date: str
    currency: str = "TRY"
    quote_id: Optional[str] = None

class SettingsIn(BaseModel):
    company_name: str = "Rhisos Mobilya"
    tagline: str = ""
    address: str = ""
    phone: str = ""
    email: str = ""
    website: str = ""
    tax_office: str = ""
    tax_number: str = ""
    logo: str = ""

class PaymentIn(BaseModel):
    amount: float
    date: str
    method: str = "cash"
    note: Optional[str] = ""

# ---------------- Serializers ----------------
def customer_dict(c: Customer) -> dict:
    return {"id": c.id, "name": c.name, "company": c.company or "", "email": c.email or "",
            "phone": c.phone or "", "address": c.address or "", "notes": c.notes or "", "created_at": c.created_at}

def transaction_dict(t: Transaction) -> dict:
    return {"id": t.id, "type": t.type, "amount": t.amount, "category": t.category,
            "payment_method": t.payment_method, "description": t.description or "", "date": t.date,
            "currency": t.currency, "quote_id": t.quote_id, "payment_id": t.payment_id,
            "auto": bool(t.auto), "created_at": t.created_at}

def quote_dict(q: Quote, items: List[QuoteItem], payments: List[Payment]) -> dict:
    return {
        "id": q.id, "quote_number": q.quote_number, "customer_id": q.customer_id,
        "customer_name": q.customer_name, "title": q.title,
        "items": [{"description": i.description, "quantity": i.quantity, "unit_price": i.unit_price, "vat_rate": i.vat_rate} for i in items],
        "currency": q.currency, "notes": q.notes or "", "valid_until": q.valid_until or "",
        "status": q.status,
        "payments": [{"id": p.id, "amount": p.amount, "date": p.date, "method": p.method, "note": p.note or ""} for p in payments],
        "paid_total": q.paid_total or 0,
        "subtotal": q.subtotal, "vat_total": q.vat_total, "discount": q.discount, "grand_total": q.grand_total,
        "created_at": q.created_at, "created_by": q.created_by,
        "emailed_at": q.emailed_at, "emailed_to": q.emailed_to,
    }

DEFAULT_SETTINGS = {
    "company_name": "Rhisos Mobilya", "tagline": "", "address": "", "phone": "",
    "email": "", "website": "", "tax_office": "", "tax_number": "", "logo": "",
}

def settings_dict(s: Settings) -> dict:
    return {
        "company_name": s.company_name or "Rhisos Mobilya", "tagline": s.tagline or "",
        "address": s.address or "", "phone": s.phone or "", "email": s.email or "",
        "website": s.website or "", "tax_office": s.tax_office or "", "tax_number": s.tax_number or "",
        "logo": s.logo or "",
    }

async def load_quote_items(db: AsyncSession, qid: str) -> List[QuoteItem]:
    res = await db.execute(select(QuoteItem).where(QuoteItem.quote_id == qid).order_by(QuoteItem.position))
    return res.scalars().all()

async def load_quote_payments(db: AsyncSession, qid: str) -> List[Payment]:
    res = await db.execute(select(Payment).where(Payment.quote_id == qid).order_by(Payment.date))
    return res.scalars().all()

def compute_quote_totals(items, discount=0):
    subtotal = 0.0
    vat_total = 0.0
    for it in items:
        line = it["quantity"] * it["unit_price"]
        subtotal += line
        vat_total += line * (it["vat_rate"] / 100.0)
    grand_total = (subtotal - discount) + vat_total
    return {"subtotal": round(subtotal, 2), "vat_total": round(vat_total, 2),
            "discount": round(discount, 2), "grand_total": round(grand_total, 2)}

# ---------------- Email (Emergent managed Resend) ----------------
EMAIL_BASE_URL = "https://integrations.emergentagent.com"
EMAIL_KEY = os.environ.get("EMERGENT_EMAIL_KEY", "")
EMAIL_FROM_NAME = os.environ.get("EMAIL_FROM_NAME", "Rhisos Mobilya")
EMAIL_REPLY_TO = os.environ.get("EMAIL_REPLY_TO")

_SHORTENERS = ("bit.ly", "tinyurl.com", "t.co", "is.gd", "cutt.ly", "goo.gl", "rebrand.ly")
_CRED_ASK = ("reply with your password", "reply with the code", "send your password", "cvv",
             "send us your password", "enter your password below", "confirm your card number",
             "your full card number", "seed phrase", "recovery phrase", "verify your card",
             "social security number", "confirm your bank details")
_HOSTISH = re.compile(r"\b(?:https?://)?((?:[a-z0-9-]+\.)+[a-z]{2,})", re.I)

def _host_ok(host: str) -> bool:
    if not host or "xn--" in host:
        return False
    try:
        ipaddress.ip_address(host)
        return False
    except ValueError:
        pass
    return not any(host == s or host.endswith("." + s) for s in _SHORTENERS)

def _same_site(shown: str, real: str) -> bool:
    return shown == real or real.endswith("." + shown) or shown.endswith("." + real)

class _EmailScan(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tags, self.urls, self.anchors = set(), [], []
        self._href, self._text = None, []
    def handle_starttag(self, tag, attrs):
        self.tags.add(tag.lower())
        self.urls += [v for k, v in attrs if k.lower() in ("href", "src") and v]
        if tag.lower() == "a":
            self._href = dict((k.lower(), v) for k, v in attrs).get("href")
            self._text = []
    def handle_data(self, data):
        if self._href is not None:
            self._text.append(data)
    def handle_endtag(self, tag):
        if tag.lower() == "a" and self._href is not None:
            self.anchors.append((self._href, "".join(self._text)))
            self._href, self._text = None, []

def _assert_safe_email(subject: str, html: str) -> None:
    scan = _EmailScan(); scan.feed(html)
    if scan.tags & {"form", "input", "textarea", "select"}:
        raise ValueError("No forms or input fields in email (G2)")
    body = f"{subject}\n{html}".lower()
    for p in _CRED_ASK:
        if p in body:
            raise ValueError(f"Email asks the recipient for credentials: {p!r} (G2)")
    for url in scan.urls:
        low = url.strip().lower()
        if low.startswith(("mailto:", "tel:", "cid:", "#")):
            continue
        if not low.startswith("https://"):
            raise ValueError(f"Email links/assets must be absolute https: {url!r} (G3)")
        host = urlparse(low).hostname or ""
        if not _host_ok(host) or urlparse(low).username is not None:
            raise ValueError(f"Shortened, numeric-host or credential-bearing URL: {url!r} (G3)")
    for href, text in scan.anchors:
        real = urlparse(href.strip().lower()).hostname or ""
        if not real:
            continue
        for m in _HOSTISH.finditer(text):
            if not _same_site(m.group(1).lower(), real):
                raise ValueError(f"Anchor text {m.group(1)!r} != real link host {real!r} (G3)")

async def send_email(*, to: str, subject: str, html: str, reply_to: str = None):
    _assert_safe_email(subject, html)
    payload = {"to": [to], "subject": subject, "html": html, "from_name": EMAIL_FROM_NAME}
    if reply_to or EMAIL_REPLY_TO:
        payload["contact_email"] = reply_to or EMAIL_REPLY_TO
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                f"{EMAIL_BASE_URL}/api/v1/email/send",
                headers={"X-Email-Key": EMAIL_KEY},
                json=payload,
            )
        resp.raise_for_status()
        return resp.json().get("id")
    except httpx.HTTPStatusError as e:
        logging.getLogger(__name__).error(f"Email send failed: {e.response.status_code} {e.response.text}")
        raise HTTPException(status_code=502, detail="E-posta gönderilemedi")
    except Exception as e:
        logging.getLogger(__name__).error(f"Email send error: {str(e)}")
        raise HTTPException(status_code=500, detail="E-posta gönderilemedi")

def _fmt_try(amount, currency="TRY"):
    sym = {"TRY": "₺", "USD": "$", "EUR": "€"}.get(currency, "")
    return f"{sym}{float(amount or 0):,.2f}"

def build_quote_email_html(doc: dict, settings_doc: dict) -> str:
    company = escape(settings_doc.get("company_name") or EMAIL_FROM_NAME)
    cur = doc.get("currency", "TRY")
    rows = ""
    for it in doc.get("items", []):
        line = it.get("quantity", 0) * it.get("unit_price", 0)
        rows += (
            f'<tr>'
            f'<td style="padding:8px 6px;border-bottom:1px solid #eee">{escape(str(it.get("description","")))}</td>'
            f'<td style="padding:8px 6px;border-bottom:1px solid #eee;text-align:right">{escape(str(it.get("quantity",0)))}</td>'
            f'<td style="padding:8px 6px;border-bottom:1px solid #eee;text-align:right">{escape(_fmt_try(it.get("unit_price",0),cur))}</td>'
            f'<td style="padding:8px 6px;border-bottom:1px solid #eee;text-align:right">%{escape(str(it.get("vat_rate",0)))}</td>'
            f'<td style="padding:8px 6px;border-bottom:1px solid #eee;text-align:right">{escape(_fmt_try(line,cur))}</td>'
            f'</tr>'
        )
    paid = doc.get("paid_total", 0)
    remaining = doc.get("grand_total", 0) - paid
    contact_bits = []
    if settings_doc.get("phone"):
        contact_bits.append(f"Tel: {escape(settings_doc['phone'])}")
    if settings_doc.get("email"):
        contact_bits.append(escape(settings_doc["email"]))
    if settings_doc.get("address"):
        contact_bits.append(escape(settings_doc["address"]))
    contact = " &nbsp;•&nbsp; ".join(contact_bits)
    return (
        f'<table role="presentation" width="100%" style="background:#FDFBF7;padding:24px 0"><tr><td align="center">'
        f'<table role="presentation" width="600" style="background:#ffffff;border:1px solid #E8E5E1;border-radius:12px;overflow:hidden;font-family:Arial,Helvetica,sans-serif;color:#1F1A17">'
        f'<tr><td style="background:#4A3B32;color:#ffffff;padding:20px 28px">'
        f'<div style="font-size:22px;font-weight:bold">{company}</div>'
        f'<div style="font-size:13px;color:#e9e2da">Teklif Belgesi</div></td></tr>'
        f'<tr><td style="padding:24px 28px">'
        f'<p style="margin:0 0 4px">Sayın <strong>{escape(str(doc.get("customer_name","")))}</strong>,</p>'
        f'<p style="margin:0 0 16px;color:#6B615A">Aşağıda <strong>{escape(str(doc.get("quote_number","")))}</strong> numaralı teklifimizin detaylarını bulabilirsiniz.</p>'
        f'<table role="presentation" width="100%" style="border-collapse:collapse;font-size:14px">'
        f'<tr style="color:#6B615A;text-align:left">'
        f'<th style="padding:8px 6px;border-bottom:2px solid #4A3B32">Açıklama</th>'
        f'<th style="padding:8px 6px;border-bottom:2px solid #4A3B32;text-align:right">Adet</th>'
        f'<th style="padding:8px 6px;border-bottom:2px solid #4A3B32;text-align:right">Birim</th>'
        f'<th style="padding:8px 6px;border-bottom:2px solid #4A3B32;text-align:right">KDV</th>'
        f'<th style="padding:8px 6px;border-bottom:2px solid #4A3B32;text-align:right">Tutar</th></tr>'
        f'{rows}</table>'
        f'<table role="presentation" width="100%" style="margin-top:16px;font-size:14px">'
        f'<tr><td style="text-align:right;color:#6B615A;padding:2px 6px">Ara Toplam</td><td style="text-align:right;padding:2px 6px;width:140px">{escape(_fmt_try(doc.get("subtotal",0),cur))}</td></tr>'
        f'<tr><td style="text-align:right;color:#6B615A;padding:2px 6px">İskonto</td><td style="text-align:right;padding:2px 6px">-{escape(_fmt_try(doc.get("discount",0),cur))}</td></tr>'
        f'<tr><td style="text-align:right;color:#6B615A;padding:2px 6px">KDV</td><td style="text-align:right;padding:2px 6px">{escape(_fmt_try(doc.get("vat_total",0),cur))}</td></tr>'
        f'<tr><td style="text-align:right;font-weight:bold;font-size:16px;padding:6px 6px;border-top:1px solid #E8E5E1">Genel Toplam</td><td style="text-align:right;font-weight:bold;font-size:16px;padding:6px 6px;border-top:1px solid #E8E5E1">{escape(_fmt_try(doc.get("grand_total",0),cur))}</td></tr>'
        f'<tr><td style="text-align:right;color:#3A5A40;padding:2px 6px">Ödenen</td><td style="text-align:right;color:#3A5A40;padding:2px 6px">{escape(_fmt_try(paid,cur))}</td></tr>'
        f'<tr><td style="text-align:right;color:#9C3D38;padding:2px 6px">Kalan Bakiye</td><td style="text-align:right;color:#9C3D38;padding:2px 6px">{escape(_fmt_try(remaining,cur))}</td></tr>'
        f'</table>'
        + (f'<p style="margin:16px 0 0;color:#6B615A;font-size:13px">{escape(str(doc.get("notes","")))}</p>' if doc.get("notes") else "")
        + f'</td></tr>'
        f'<tr><td style="padding:16px 28px;background:#FDFBF7;color:#6B615A;font-size:12px;border-top:1px solid #E8E5E1">'
        f'{contact}<br/>Bu e-posta {company} tarafından gönderilmiştir. Şifre veya kart bilgisi asla e-posta ile istenmez.</td></tr>'
        f'</table></td></tr></table>'
    )

# ---------------- Auth Routes ----------------
@api_router.post("/auth/register")
async def register(body: RegisterIn, response: Response, db: AsyncSession = Depends(get_db)):
    email = body.email.lower()
    existing = await db.execute(select(User).where(User.email == email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Bu email zaten kayıtlı")
    user = User(id=new_id(), email=email, password_hash=hash_password(body.password),
                name=body.name, role="user", created_at=now_iso())
    db.add(user)
    await db.commit()
    set_auth_cookies(response, create_access_token(user.id, email), create_refresh_token(user.id))
    return {"id": user.id, "email": email, "name": body.name, "role": "user"}

@api_router.post("/auth/login")
async def login(body: LoginIn, request: Request, response: Response, db: AsyncSession = Depends(get_db)):
    email = body.email.lower()
    ip = request.client.host if request.client else "unknown"
    ident = f"{ip}:{email}"
    attempt = await db.get(LoginAttempt, ident)
    if attempt and (attempt.count or 0) >= 5 and attempt.locked_until:
        if datetime.fromisoformat(attempt.locked_until) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Çok fazla başarısız deneme. 15 dakika sonra tekrar deneyin.")
    res = await db.execute(select(User).where(User.email == email))
    user = res.scalar_one_or_none()
    if not user or not verify_password(body.password, user.password_hash):
        locked = (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()
        if attempt:
            attempt.count = (attempt.count or 0) + 1
            attempt.locked_until = locked
        else:
            db.add(LoginAttempt(identifier=ident, count=1, locked_until=locked))
        await db.commit()
        raise HTTPException(status_code=401, detail="Email veya şifre hatalı")
    if attempt:
        await db.delete(attempt)
        await db.commit()
    set_auth_cookies(response, create_access_token(user.id, email), create_refresh_token(user.id))
    return {"id": user.id, "email": email, "name": user.name, "role": user.role or "user"}

@api_router.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}

@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

@api_router.post("/auth/refresh")
async def refresh(request: Request, response: Response, db: AsyncSession = Depends(get_db)):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Refresh token yok")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Geçersiz token")
        user = await db.get(User, payload["sub"])
        if not user:
            raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
        access = create_access_token(user.id, user.email)
        response.set_cookie("access_token", access, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
        return {"ok": True}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz token")

# ---------------- Customer Routes ----------------
@api_router.get("/customers")
async def list_customers(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Customer).order_by(Customer.created_at.desc()))
    return [customer_dict(c) for c in res.scalars().all()]

@api_router.post("/customers")
async def create_customer(body: CustomerIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    c = Customer(id=new_id(), created_at=now_iso(), **body.model_dump())
    db.add(c)
    await db.commit()
    return customer_dict(c)

@api_router.get("/customers/{cid}")
async def get_customer(cid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    c = await db.get(Customer, cid)
    if not c:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    return customer_dict(c)

@api_router.put("/customers/{cid}")
async def update_customer(cid: str, body: CustomerIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    c = await db.get(Customer, cid)
    if not c:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    for k, v in body.model_dump().items():
        setattr(c, k, v)
    await db.commit()
    return customer_dict(c)

@api_router.delete("/customers/{cid}")
async def delete_customer(cid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    c = await db.get(Customer, cid)
    if c:
        await db.delete(c)
        await db.commit()
    return {"ok": True}

@api_router.get("/customers/{cid}/history")
async def customer_history(cid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    customer = await db.get(Customer, cid)
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    res = await db.execute(select(Quote).where(Quote.customer_id == cid).order_by(Quote.created_at.desc()))
    quotes = res.scalars().all()
    quote_dicts = []
    payments = []
    for q in quotes:
        items = await load_quote_items(db, q.id)
        pays = await load_quote_payments(db, q.id)
        quote_dicts.append(quote_dict(q, items, pays))
        for p in pays:
            payments.append({"id": p.id, "amount": p.amount, "date": p.date, "method": p.method,
                             "note": p.note or "", "quote_id": q.id, "quote_number": q.quote_number, "currency": q.currency})
    payments.sort(key=lambda x: x.get("date", ""), reverse=True)
    total_quoted = round(sum(q.grand_total or 0 for q in quotes), 2)
    total_approved = round(sum(q.grand_total or 0 for q in quotes if q.status == "approved"), 2)
    total_paid = round(sum(p["amount"] for p in payments), 2)
    return {
        "customer": customer_dict(customer),
        "quotes": quote_dicts,
        "payments": payments,
        "totals": {"total_quoted": total_quoted, "total_approved": total_approved,
                   "total_paid": total_paid, "quote_count": len(quotes)},
    }

# ---------------- Quote Routes ----------------
async def next_quote_number(db: AsyncSession) -> str:
    res = await db.execute(select(func.count(Quote.id)))
    count = res.scalar_one()
    return f"TKF-{datetime.now().year}-{count + 1:04d}"

@api_router.get("/quotes")
async def list_quotes(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Quote).order_by(Quote.created_at.desc()))
    quotes = res.scalars().all()
    out = []
    for q in quotes:
        items = await load_quote_items(db, q.id)
        pays = await load_quote_payments(db, q.id)
        out.append(quote_dict(q, items, pays))
    return out

@api_router.get("/quotes/{qid}")
async def get_quote(qid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if not q:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    items = await load_quote_items(db, qid)
    pays = await load_quote_payments(db, qid)
    return quote_dict(q, items, pays)

@api_router.post("/quotes")
async def create_quote(body: QuoteIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    items = [i.model_dump() for i in body.items]
    totals = compute_quote_totals(items, body.discount)
    qid = new_id()
    q = Quote(
        id=qid, quote_number=await next_quote_number(db), customer_id=body.customer_id,
        customer_name=body.customer_name, title=body.title, currency=body.currency,
        notes=body.notes, valid_until=body.valid_until, status="pending", paid_total=0,
        created_at=now_iso(), created_by=user["name"], **totals,
    )
    db.add(q)
    for idx, it in enumerate(items):
        db.add(QuoteItem(id=new_id(), quote_id=qid, position=idx, **it))
    await db.commit()
    item_objs = await load_quote_items(db, qid)
    return quote_dict(q, item_objs, [])

@api_router.put("/quotes/{qid}")
async def update_quote(qid: str, body: QuoteIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if not q:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    items = [i.model_dump() for i in body.items]
    totals = compute_quote_totals(items, body.discount)
    q.customer_id = body.customer_id
    q.customer_name = body.customer_name
    q.title = body.title
    q.currency = body.currency
    q.notes = body.notes
    q.valid_until = body.valid_until
    q.subtotal = totals["subtotal"]
    q.vat_total = totals["vat_total"]
    q.discount = totals["discount"]
    q.grand_total = totals["grand_total"]
    await db.execute(sa_delete(QuoteItem).where(QuoteItem.quote_id == qid))
    for idx, it in enumerate(items):
        db.add(QuoteItem(id=new_id(), quote_id=qid, position=idx, **it))
    await db.commit()
    item_objs = await load_quote_items(db, qid)
    pays = await load_quote_payments(db, qid)
    return quote_dict(q, item_objs, pays)

@api_router.patch("/quotes/{qid}/status")
async def update_quote_status(qid: str, body: QuoteStatusIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if not q:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    q.status = body.status
    await db.commit()
    items = await load_quote_items(db, qid)
    pays = await load_quote_payments(db, qid)
    return quote_dict(q, items, pays)

@api_router.post("/quotes/{qid}/payments")
async def add_quote_payment(qid: str, body: PaymentIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if not q:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="Geçerli bir tutar girin")
    pid = new_id()
    db.add(Payment(id=pid, quote_id=qid, amount=round(body.amount, 2), date=body.date, method=body.method, note=body.note or ""))
    await db.flush()
    pays = await load_quote_payments(db, qid)
    q.paid_total = round(sum(p.amount for p in pays), 2)
    db.add(Transaction(
        id=new_id(), type="income", amount=round(body.amount, 2), category="Kapora / Teklif Ödemesi",
        payment_method=body.method, description=f"{q.quote_number} - {q.customer_name} ödemesi",
        date=body.date, currency=q.currency, quote_id=qid, payment_id=pid, auto=True, created_at=now_iso(),
    ))
    await db.commit()
    items = await load_quote_items(db, qid)
    pays = await load_quote_payments(db, qid)
    return quote_dict(q, items, pays)

@api_router.delete("/quotes/{qid}/payments/{pid}")
async def delete_quote_payment(qid: str, pid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if not q:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    p = await db.get(Payment, pid)
    if p:
        await db.delete(p)
    await db.execute(sa_delete(Transaction).where(Transaction.payment_id == pid))
    await db.flush()
    pays = await load_quote_payments(db, qid)
    q.paid_total = round(sum(x.amount for x in pays), 2)
    await db.commit()
    items = await load_quote_items(db, qid)
    pays = await load_quote_payments(db, qid)
    return quote_dict(q, items, pays)

@api_router.post("/quotes/{qid}/email")
async def email_quote(qid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if not q:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    customer = await db.get(Customer, q.customer_id)
    to_email = ((customer.email if customer else "") or "").strip()
    if not to_email:
        raise HTTPException(status_code=400, detail="Müşterinin kayıtlı e-posta adresi yok. Önce müşteriye e-posta ekleyin.")
    settings_row = await db.get(Settings, "company")
    settings_data = settings_dict(settings_row) if settings_row else DEFAULT_SETTINGS
    items = await load_quote_items(db, qid)
    pays = await load_quote_payments(db, qid)
    doc = quote_dict(q, items, pays)
    subject = f"{EMAIL_FROM_NAME} - Teklif {q.quote_number}"
    html = build_quote_email_html(doc, settings_data)
    email_id = await send_email(to=to_email, subject=subject, html=html)
    q.emailed_at = now_iso()
    q.emailed_to = to_email
    await db.commit()
    return {"ok": True, "email_id": email_id, "to": to_email}

@api_router.delete("/quotes/{qid}")
async def delete_quote(qid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if q:
        await db.delete(q)
    await db.execute(sa_delete(QuoteItem).where(QuoteItem.quote_id == qid))
    await db.execute(sa_delete(Payment).where(Payment.quote_id == qid))
    await db.execute(sa_delete(Transaction).where(Transaction.quote_id == qid))
    await db.commit()
    return {"ok": True}

# ---------------- Public (no auth) ----------------
@api_router.get("/public/quotes/{qid}")
async def public_quote(qid: str, db: AsyncSession = Depends(get_db)):
    q = await db.get(Quote, qid)
    if not q:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    items = await load_quote_items(db, qid)
    pays = await load_quote_payments(db, qid)
    settings_row = await db.get(Settings, "company")
    company = {**DEFAULT_SETTINGS, **(settings_dict(settings_row) if settings_row else {})}
    return {"quote": quote_dict(q, items, pays), "company": company}

# ---------------- Transaction Routes ----------------
@api_router.get("/transactions/export")
async def export_transactions(type: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None,
                              user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    stmt = select(Transaction)
    if type:
        stmt = stmt.where(Transaction.type == type)
    if start:
        stmt = stmt.where(Transaction.date >= start)
    if end:
        stmt = stmt.where(Transaction.date <= end)
    stmt = stmt.order_by(Transaction.date)
    res = await db.execute(stmt)
    docs = res.scalars().all()

    type_tr = {"income": "Gelir", "expense": "Gider"}
    pay_tr = {"cash": "Nakit", "bank": "Banka/Havale", "card": "Kredi Kartı", "check": "Çek/Senet"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kasa Raporu"
    headers = ["Tarih", "Tür", "Kategori", "Ödeme Yöntemi", "Açıklama", "Tutar", "Para Birimi"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4A3B32", end_color="4A3B32", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_income = 0.0
    total_expense = 0.0
    for d in docs:
        if d.type == "income":
            total_income += d.amount
        else:
            total_expense += d.amount
        ws.append([d.date, type_tr.get(d.type, d.type), d.category, pay_tr.get(d.payment_method, d.payment_method),
                   d.description or "", d.amount, d.currency])

    ws.append([])
    ws.append(["", "", "", "", "Toplam Gelir", round(total_income, 2), ""])
    ws.append(["", "", "", "", "Toplam Gider", round(total_expense, 2), ""])
    ws.append(["", "", "", "", "Net Bakiye", round(total_income - total_expense, 2), ""])
    for row in ws.iter_rows(min_row=ws.max_row - 2, max_row=ws.max_row):
        row[4].font = Font(bold=True)
        row[5].font = Font(bold=True)

    widths = [14, 10, 20, 16, 40, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rhisos_kasa_raporu.xlsx"},
    )

@api_router.get("/transactions")
async def list_transactions(type: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None,
                            category: Optional[str] = None, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    stmt = select(Transaction)
    if type:
        stmt = stmt.where(Transaction.type == type)
    if category:
        stmt = stmt.where(Transaction.category == category)
    if start:
        stmt = stmt.where(Transaction.date >= start)
    if end:
        stmt = stmt.where(Transaction.date <= end)
    stmt = stmt.order_by(Transaction.date.desc())
    res = await db.execute(stmt)
    return [transaction_dict(t) for t in res.scalars().all()]

@api_router.post("/transactions")
async def create_transaction(body: TransactionIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    t = Transaction(id=new_id(), auto=False, created_at=now_iso(), **body.model_dump())
    db.add(t)
    await db.commit()
    return transaction_dict(t)

@api_router.put("/transactions/{tid}")
async def update_transaction(tid: str, body: TransactionIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    t = await db.get(Transaction, tid)
    if not t:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    for k, v in body.model_dump().items():
        setattr(t, k, v)
    await db.commit()
    return transaction_dict(t)

@api_router.delete("/transactions/{tid}")
async def delete_transaction(tid: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    t = await db.get(Transaction, tid)
    if t:
        await db.delete(t)
        await db.commit()
    return {"ok": True}

# ---------------- Dashboard ----------------
@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    txns = (await db.execute(select(Transaction))).scalars().all()
    total_income = sum(t.amount for t in txns if t.type == "income")
    total_expense = sum(t.amount for t in txns if t.type == "expense")

    monthly = {}
    for t in txns:
        month = (t.date or "")[:7]
        if not month:
            continue
        monthly.setdefault(month, {"month": month, "income": 0, "expense": 0})
        monthly[month][t.type] += t.amount
    monthly_list = sorted(monthly.values(), key=lambda x: x["month"])[-6:]

    cat = {}
    for t in txns:
        if t.type == "expense":
            cat[t.category] = cat.get(t.category, 0) + t.amount
    expense_by_cat = [{"name": k, "value": round(v, 2)} for k, v in cat.items()]

    quotes = (await db.execute(select(Quote))).scalars().all()
    quote_counts = {"pending": 0, "approved": 0, "rejected": 0}
    expiring = []
    today = datetime.now(timezone.utc).date()
    soon = today + timedelta(days=7)
    for q in quotes:
        st = q.status or "pending"
        quote_counts[st] = quote_counts.get(st, 0) + 1
        vu = q.valid_until
        if st == "pending" and vu:
            try:
                vu_date = date.fromisoformat(vu)
            except (ValueError, TypeError):
                continue
            if vu_date <= soon:
                expiring.append({"id": q.id, "quote_number": q.quote_number, "customer_name": q.customer_name,
                                 "valid_until": vu, "days_left": (vu_date - today).days,
                                 "grand_total": q.grand_total or 0, "currency": q.currency or "TRY"})
    expiring.sort(key=lambda x: x["days_left"])

    customer_count = (await db.execute(select(func.count(Customer.id)))).scalar_one()

    return {
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "balance": round(total_income - total_expense, 2),
        "monthly": monthly_list,
        "expense_by_category": expense_by_cat,
        "quote_counts": quote_counts,
        "total_quotes": len(quotes),
        "customer_count": customer_count,
        "expiring_quotes": expiring,
    }

@api_router.get("/dashboard/monthly-profit")
async def monthly_profit(month: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    stmt = select(Transaction).where(Transaction.date >= f"{month}-01", Transaction.date <= f"{month}-31")
    txns = (await db.execute(stmt)).scalars().all()
    income = round(sum(t.amount for t in txns if t.type == "income"), 2)
    expense = round(sum(t.amount for t in txns if t.type == "expense"), 2)
    return {"month": month, "income": income, "expense": expense, "profit": round(income - expense, 2)}

# ---------------- Settings ----------------
@api_router.get("/settings")
async def get_settings(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    s = await db.get(Settings, "company")
    if not s:
        return DEFAULT_SETTINGS
    return settings_dict(s)

@api_router.put("/settings")
async def update_settings(body: SettingsIn, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    s = await db.get(Settings, "company")
    data = body.model_dump()
    if s:
        for k, v in data.items():
            setattr(s, k, v)
    else:
        db.add(Settings(key="company", **data))
    await db.commit()
    return data

# ---------------- App wiring ----------------
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get("FRONTEND_URL", "http://localhost:3000")],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    async with AsyncSessionLocal() as db:
        res = await db.execute(select(User).where(User.email == admin_email))
        existing = res.scalar_one_or_none()
        if existing is None:
            db.add(User(id=new_id(), email=admin_email, password_hash=hash_password(admin_password),
                        name="Rhisos Admin", role="admin", created_at=now_iso()))
            await db.commit()
            logger.info("Admin seeded: %s", admin_email)
        elif not verify_password(admin_password, existing.password_hash):
            existing.password_hash = hash_password(admin_password)
            await db.commit()

@app.on_event("shutdown")
async def shutdown():
    await engine.dispose()
